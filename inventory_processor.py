# -*- coding: utf-8 -*-
"""出入库流水清洗处理器 — 纯逻辑模块，不含 GUI"""

import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============== 配置常量 ==============
DATA_START_ROW = 8  # 0索引：第9行开始是数据

DEPARTMENT_MAP = {
    '彭建森': 'VCP',
    '罗德浮': 'VCP',
    '曹鹏': '蚀刻',
    '邹柳珍': '曝光',
    '刘志军': '品质',
}

KEEP_ORDER_TYPE = '材料出库单'

OUTPUT_COLUMNS = [
    '单据日期', '单据类型', '经手人', '部门', '仓库',
    '存货', '规格型号', '计量单位',
    '出库数量', '出库单价', '出库金额',
]

NORMALIZED_COLUMNS = [
    '单据日期', '单据类型', '部门', '经手人', '往来单位',
    '仓库', '存货', '规格型号', '计量单位',
    '入库数量', '入库单价', '入库金额',
    '出库数量', '出库单价', '出库金额',
]

DEPT_SHEET_ORDER = ['VCP', '曝光', '蚀刻', '品质', '其他']


def _safe_to_numeric(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0
    s = re.sub(r'[¥$￥,\s]', '', s)
    try:
        return float(s)
    except ValueError:
        return 0


def _is_data_row(row_values):
    if row_values is None or len(row_values) < 2:
        return False
    col1 = row_values[1] if len(row_values) > 1 else None
    col2 = row_values[2] if len(row_values) > 2 else None
    if col1 is None or (isinstance(col1, float) and pd.isna(col1)):
        return False
    s = str(col1).strip()
    if not s or '合计' in s or '制表' in s or '打印' in s or '第' in s:
        return False
    if re.match(r'\d{4}[-/.]\d{1,2}[-/.]\d{1,2}', s):
        return True
    if col2 is not None:
        col2_s = str(col2).strip()
        if col2_s in [KEEP_ORDER_TYPE, '进货单', '销货单']:
            return True
    return False


def process_inventory(input_path, output_path=None, progress_callback=None):
    """主处理函数

    Args:
        input_path: 输入 Excel 文件路径
        output_path: 输出 Excel 文件路径（可选，默认自动生成）
        progress_callback: 进度回调 (current_step, total_steps, message)

    Returns:
        dict: 处理结果摘要
    """
    def prog(current, total, message):
        if progress_callback:
            progress_callback(current, total, message)

    if not output_path:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(
            os.path.dirname(input_path) or '.',
            f'出入库流水_清洗后_{ts}.xlsx'
        )

    # ---- Step 1: 读取 ----
    prog(1, 6, "读取文件...")
    df_raw = pd.read_excel(input_path, header=None, dtype=object)
    if df_raw.empty:
        raise ValueError("文件为空，无法处理")

    data_rows = []
    for r in range(DATA_START_ROW, len(df_raw)):
        row_vals = list(df_raw.iloc[r])
        if _is_data_row(row_vals):
            data_rows.append(row_vals)

    data_rows_trimmed = [row[1:] for row in data_rows]
    df = pd.DataFrame(data_rows_trimmed, columns=NORMALIZED_COLUMNS)

    # ---- Step 2: 筛选 ----
    prog(2, 6, "筛选材料出库单...")
    if '单据类型' not in df.columns:
        raise KeyError(f"找不到'单据类型'列，可用列: {list(df.columns)}")
    before = len(df)
    df = df[df['单据类型'] == KEEP_ORDER_TYPE].copy()
    after = len(df)

    # ---- Step 3: 清洗数值 ----
    prog(3, 6, "清洗数值列...")
    for col in ['出库数量', '出库单价', '出库金额']:
        if col in df.columns:
            df[col] = df[col].apply(_safe_to_numeric)

    # ---- Step 4: 部门映射 ----
    prog(4, 6, "部门映射...")
    if '经手人' not in df.columns:
        raise KeyError(f"找不到'经手人'列，可用列: {list(df.columns)}")

    new_depts = []
    for _, row in df.iterrows():
        person = row['经手人']
        if pd.isna(person):
            person = ''
        person = str(person).strip()
        new_depts.append(DEPARTMENT_MAP.get(person, '其他'))
    df['部门'] = new_depts
    mapped = len([d for d in new_depts if d != '其他'])

    # ---- Step 5: 排序汇总 ----
    prog(5, 6, "排序与汇总...")
    df = df.sort_values(by=['存货', '部门'], ascending=[True, True]).reset_index(drop=True)
    summary = df.groupby('存货').agg(
        出库数量合计=('出库数量', 'sum'),
        出库金额合计=('出库金额', 'sum'),
        记录条数=('出库数量', 'count'),
    ).reset_index()
    summary = summary.sort_values(by='存货', ascending=True).reset_index(drop=True)
    summary['出库数量合计'] = summary['出库数量合计'].round(2)
    summary['出库金额合计'] = summary['出库金额合计'].round(2)

    # ---- Step 6: 生成 Excel ----
    prog(6, 6, "生成 Excel 报告...")
    _generate_excel(df, summary, output_path)

    # 部门统计
    dept_stats = df.groupby('部门').agg(
        记录数=('出库金额', 'count'),
        出库金额合计=('出库金额', 'sum'),
        出库数量合计=('出库数量', 'sum'),
    ).round(2)

    dept_info = {}
    for dn in DEPT_SHEET_ORDER:
        if dn in dept_stats.index:
            r = dept_stats.loc[dn]
            dept_info[dn] = {
                'count': int(r['记录数']),
                'qty': round(r['出库数量合计'], 2),
                'amount': round(r['出库金额合计'], 2),
            }
        else:
            dept_info[dn] = {'count': 0, 'qty': 0, 'amount': 0}

    return {
        'input_path': input_path,
        'output_path': output_path,
        'total_records': len(df),
        'inventory_count': len(summary),
        'total_qty': round(summary['出库数量合计'].sum(), 2),
        'total_amount': round(summary['出库金额合计'].sum(), 2),
        'mapped_count': mapped,
        'other_count': len(df) - mapped,
        'raw_count': before,
        'filtered_count': after,
        'dept_info': dept_info,
    }


def _generate_excel(df, summary, output_path):
    """生成多工作表 Excel 输出文件"""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    data_font = Font(name='微软雅黑', size=10)
    total_font = Font(name='微软雅黑', bold=True, size=10)
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def style_data_cell(ws, row, col, align='center'):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border
        if align == 'right':
            cell.alignment = right_align
        elif align == 'left':
            cell.alignment = left_align
        else:
            cell.alignment = center_align

    def auto_width(ws, min_width=8, max_width=40):
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    val = str(cell.value)
                    char_len = sum(2 if '一' <= ch <= '鿿' or '　' <= ch <= '〿' or '＀' <= ch <= '￯' else 1 for ch in val)
                    max_len = max(max_len, char_len)
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))

    # ---- 汇总表 ----
    ws_summary = wb.create_sheet('汇总')
    summary_cols = ['存货', '出库金额合计']
    ws_summary.append(summary_cols)
    style_header(ws_summary, len(summary_cols))

    for _, row_data in summary.iterrows():
        ws_summary.append([row_data['存货'], round(row_data['出库金额合计'], 2)])

    total_row = len(summary) + 2
    total_amount = round(summary['出库金额合计'].sum(), 2)
    ws_summary.append(['合计', total_amount])
    for col in range(1, len(summary_cols) + 1):
        cell = ws_summary.cell(row=total_row, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = center_align if col == 1 else right_align

    for r in range(2, total_row):
        style_data_cell(ws_summary, r, 1, 'left')
        style_data_cell(ws_summary, r, 2, 'right')
    auto_width(ws_summary)
    ws_summary.freeze_panes = 'A2'

    # ---- 分部门明细 ----
    available_cols = [c for c in OUTPUT_COLUMNS if c in df.columns]
    summary_header_col = 14  # N列
    dept_summary_cols = ['存货', '总数量', '平均单价', '总出库金额']

    for dept_name in DEPT_SHEET_ORDER:
        ws_dept = wb.create_sheet(dept_name)
        dept_df = df[df['部门'] == dept_name].copy()
        dept_df = dept_df.sort_values(by='存货', ascending=True)

        ws_dept.append(available_cols)
        style_header(ws_dept, len(available_cols))

        if len(dept_df) == 0:
            ws_dept.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(available_cols))
            ws_dept.cell(row=2, column=1, value='无数据')
            ws_dept.cell(row=2, column=1).font = Font(name='微软雅黑', size=10, color='999999')
            ws_dept.cell(row=2, column=1).alignment = center_align
            for ci, cn in enumerate(dept_summary_cols):
                c = summary_header_col + ci
                ws_dept.cell(row=1, column=c, value=cn)
                cell = ws_dept.cell(row=1, column=c)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            ws_dept.merge_cells(start_row=2, start_column=summary_header_col,
                                end_row=2, end_column=summary_header_col + len(dept_summary_cols) - 1)
            ws_dept.cell(row=2, column=summary_header_col, value='无数据')
            ws_dept.cell(row=2, column=summary_header_col).font = Font(name='微软雅黑', size=10, color='999999')
            ws_dept.cell(row=2, column=summary_header_col).alignment = center_align
            auto_width(ws_dept)
            ws_dept.freeze_panes = 'A2'
            continue

        for r_idx, (_, row_data) in enumerate(dept_df.iterrows()):
            row_num = r_idx + 2
            for c_idx, col_name in enumerate(available_cols):
                val = row_data[col_name]
                if isinstance(val, float) and pd.isna(val):
                    val = ''
                ws_dept.cell(row=row_num, column=c_idx + 1, value=val)

        for r in range(2, len(dept_df) + 2):
            for c in range(1, len(available_cols) + 1):
                col_name = available_cols[c - 1]
                style_data_cell(ws_dept, r, c, 'right' if col_name in ['出库数量', '出库单价', '出库金额'] else 'center')

        # 底部汇总行
        summary_row = len(dept_df) + 2
        qty_sum = round(dept_df['出库数量'].sum(), 2)
        amt_sum = round(dept_df['出库金额'].sum(), 2)

        if '存货' in available_cols:
            ws_dept.cell(row=summary_row, column=available_cols.index('存货') + 1, value='汇总')
        else:
            ws_dept.cell(row=summary_row, column=1, value='汇总')
        if '出库数量' in available_cols:
            ws_dept.cell(row=summary_row, column=available_cols.index('出库数量') + 1, value=qty_sum)
        if '出库金额' in available_cols:
            ws_dept.cell(row=summary_row, column=available_cols.index('出库金额') + 1, value=amt_sum)

        for c in range(1, len(available_cols) + 1):
            cell = ws_dept.cell(row=summary_row, column=c)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            cell.alignment = center_align

        # 右侧按存货汇总 (N列开始)
        for ci, cn in enumerate(dept_summary_cols):
            c = summary_header_col + ci
            ws_dept.cell(row=1, column=c, value=cn)
            cell = ws_dept.cell(row=1, column=c)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        dept_inv = dept_df.groupby('存货').agg(
            总数量=('出库数量', 'sum'),
            总出库金额=('出库金额', 'sum'),
        ).reset_index()
        dept_inv['平均单价'] = (dept_inv['总出库金额'] / dept_inv['总数量'].replace(0, None)).round(2)
        dept_inv['总数量'] = dept_inv['总数量'].round(2)
        dept_inv['总出库金额'] = dept_inv['总出库金额'].round(2)
        dept_inv = dept_inv.sort_values(by='存货', ascending=True)

        for ri, (_, inv_row) in enumerate(dept_inv.iterrows()):
            row_num = ri + 2
            vals = [inv_row['存货'], inv_row['总数量'], inv_row['平均单价'], inv_row['总出库金额']]
            for ci, val in enumerate(vals):
                c = summary_header_col + ci
                ws_dept.cell(row=row_num, column=c, value=val)
                cell = ws_dept.cell(row=row_num, column=c)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = right_align if ci > 0 else left_align

        inv_total_row = len(dept_inv) + 2
        inv_total_qty = round(dept_inv['总数量'].sum(), 2)
        inv_total_amt = round(dept_inv['总出库金额'].sum(), 2)
        inv_total_avg = round(inv_total_amt / inv_total_qty, 2) if inv_total_qty != 0 else 0
        inv_total_vals = ['合计', inv_total_qty, inv_total_avg, inv_total_amt]
        for ci, val in enumerate(inv_total_vals):
            c = summary_header_col + ci
            ws_dept.cell(row=inv_total_row, column=c, value=val)
            cell = ws_dept.cell(row=inv_total_row, column=c)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            cell.alignment = center_align if ci == 0 else right_align

        auto_width(ws_dept)
        ws_dept.freeze_panes = 'A2'

    wb.save(output_path)


# 命令行入口
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='出入库流水自动化清洗')
    parser.add_argument('--input', '-i', default='出入库流水.xlsx', help='输入Excel文件')
    parser.add_argument('--output', '-o', default=None, help='输出Excel文件')
    args = parser.parse_args()

    result = process_inventory(args.input, args.output)
    print("处理完成！")
    print(f"  总记录数: {result['total_records']}")
    print(f"  存货种类: {result['inventory_count']}")
    print(f"  出库金额合计: {result['total_amount']:,.2f}")
    print(f"  输出: {result['output_path']}")
