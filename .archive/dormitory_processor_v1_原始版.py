# -*- coding: utf-8 -*-
"""宿舍水电住宿费分摊 — 纯逻辑模块（无 UI 依赖）"""

import os
import re
from datetime import datetime, timedelta
from calendar import monthrange

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def parse_billing_month_from_filename(filepath):
    basename = os.path.splitext(os.path.basename(filepath))[0]
    m = re.search(r'(\d{4})-(\d{1,2})$', basename)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def _safe_str(val):
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _calc_effective_days(checkin, checkout, month_start, month_end):
    if checkin is None or pd.isna(checkin):
        return 0
    if checkout is None or pd.isna(checkout):
        checkout = month_end
    if isinstance(checkin, datetime):
        checkin = checkin.date()
    if isinstance(checkout, datetime):
        checkout = checkout.date()
    if isinstance(month_start, datetime):
        month_start = month_start.date()
    if isinstance(month_end, datetime):
        month_end = month_end.date()
    effective_start = max(checkin, month_start)
    effective_end = min(checkout, month_end)
    days = (effective_end - effective_start).days + 1
    return max(0, days)


def _forward_fill_room_info(df):
    for col in ["房间号码", "房型", "入住人数"]:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: _safe_str(x) if pd.notna(x) else None)
            df[col] = df[col].ffill()
    return df


def _read_utility_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["水电总计表"]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append(row)
    if len(rows) < 2:
        wb.close()
        return {}, "room"
    header = str(rows[0][0]) if rows[0][0] else ""
    mode = "room"
    if "月份" in header or (isinstance(rows[1][0], str) and re.match(r"\d{4}-\d{2}", str(rows[1][0]))):
        mode = "month"
    result = {}
    for i, row in enumerate(rows[1:], start=2):
        key = row[0]
        val = row[1] if len(row) > 1 else 0
        if key is None or (isinstance(key, str) and key.strip() in ("", "总计", "合计")):
            continue
        try:
            val = float(val) if val is not None else 0.0
        except (ValueError, TypeError):
            continue
        if mode == "month":
            result[str(key).strip()] = val
        else:
            result[_safe_str(key)] = val
    wb.close()
    return result, mode


def _read_housing_sheet(filepath):
    df = pd.read_excel(filepath, sheet_name="房屋表")
    col_aliases = {
        "房间号码": ["房间号码", "单元房号", "房号"],
        "房型": ["房型"],
        "入住人数": ["入住人数"],
        "入住人员": ["入住人员", "姓名"],
        "住宿计费时间": ["住宿计费时间", "入住日期", "计费开始"],
        "截止日期": ["截止日期", "离开日期", "计费结束"],
    }
    rename_map = {}
    for target, candidates in col_aliases.items():
        for c in candidates:
            if c in df.columns:
                rename_map[c] = target
                break
    df = df.rename(columns=rename_map)
    df = _forward_fill_room_info(df)
    if "入住人员" in df.columns:
        df = df[df["入住人员"].notna()].copy()
        df = df[df["入住人员"].astype(str).str.strip() != ""].copy()
    for col in ["住宿计费时间", "截止日期"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    if "入住人数" in df.columns:
        df["入住人数"] = pd.to_numeric(df["入住人数"], errors="coerce")
    return df


def _write_excel(output_path, df_person, df_room):
    wb = openpyxl.Workbook()
    hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="微软雅黑", size=10)
    calign = Alignment(horizontal="center", vertical="center")
    tb = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    hb = Border(
        left=Side(style="thin", color="2F5496"),
        right=Side(style="thin", color="2F5496"),
        top=Side(style="thin", color="2F5496"),
        bottom=Side(style="thin", color="2F5496"),
    )

    def write_sheet(ws, df, money_cols):
        for ci, cn in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=ci, value=cn)
            c.font = hf; c.fill = hfill; c.alignment = halign; c.border = hb
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, cn in enumerate(df.columns, 1):
                v = row[cn]
                if pd.isna(v):
                    v = ""
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = cf; c.alignment = calign; c.border = tb
                if cn in money_cols and isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
        for ci in range(1, len(df.columns) + 1):
            ml = 0
            for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=ci, max_col=ci):
                for c in r:
                    if c.value:
                        s = str(c.value)
                        ml = max(ml, sum(2 if ord(ch) > 127 else 1 for ch in s))
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 4, 36)
        ws.freeze_panes = "A2"

    ws1 = wb.active; ws1.title = "按个人明细"
    write_sheet(ws1, df_person, ["水电分摊金额(元)", "住宿费(元)", "总应付金额(元)"])
    ws2 = wb.create_sheet("按房间汇总")
    write_sheet(ws2, df_room, ["房间总应付金额(元)"])
    wb.save(output_path)


def process_allocation(filepath, billing_year, billing_month,
                       treat_end_of_prev_month_as_living=False,
                       progress_callback=None):
    utility_data, utility_mode = _read_utility_sheet(filepath)
    housing_df = _read_housing_sheet(filepath)

    month_start = datetime(billing_year, billing_month, 1)
    _, last_day = monthrange(billing_year, billing_month)
    month_end = datetime(billing_year, billing_month, last_day)
    total_days_in_month = last_day
    prev_month_last_day = month_start - timedelta(days=1)

    if treat_end_of_prev_month_as_living:
        count_fixed = 0
        for idx in housing_df.index:
            d = housing_df.at[idx, "截止日期"]
            if pd.notna(d):
                d_date = d.date() if isinstance(d, datetime) else d
                if d_date == prev_month_last_day.date():
                    housing_df.at[idx, "截止日期"] = pd.NaT
                    count_fixed += 1

    results_person = []
    results_room = {}
    all_rooms = set()
    if utility_mode == "room":
        all_rooms.update(utility_data.keys())
    all_rooms.update(housing_df["房间号码"].dropna().apply(_safe_str).unique())
    errors = []

    for room in sorted(all_rooms):
        room_residents = housing_df[housing_df["房间号码"].apply(_safe_str) == room]
        if len(room_residents) == 0:
            if utility_mode == "room" and room in utility_data:
                errors.append(f"房间 {room}: 有账单但无入住人员")
            continue
        room_utility = utility_data.get(room, 0.0) if utility_mode == "room" else 0.0
        resident_days = []
        for idx, row in room_residents.iterrows():
            try:
                days = _calc_effective_days(
                    row["住宿计费时间"], row["截止日期"], month_start, month_end)
            except Exception:
                continue
            resident_days.append((row, days))
        total_person_days = sum(d for _, d in resident_days)
        for row_data, days in resident_days:
            person_name = str(row_data["入住人员"]) if pd.notna(row_data["入住人员"]) else ""
            if total_person_days > 0 and room_utility > 0:
                utility_share = room_utility * (days / total_person_days)
            else:
                utility_share = 0.0
            accommodation = 50.0 * (days / total_days_in_month) if days > 0 else 0.0
            results_person.append({
                "月份": f"{billing_year}-{billing_month:02d}",
                "房间号码": room,
                "入住人员": person_name,
                "有效居住天数": days,
                "水电分摊金额(元)": round(utility_share, 2),
                "住宿费(元)": round(accommodation, 2),
                "总应付金额(元)": round(utility_share + accommodation, 2),
            })
            mk = f"{billing_year}-{billing_month:02d}"
            if room not in results_room:
                results_room[room] = {}
            if mk not in results_room[room]:
                results_room[room][mk] = 0.0
            results_room[room][mk] += round(utility_share + accommodation, 2)

    df_person = pd.DataFrame(results_person)
    if len(df_person) == 0:
        df_person = pd.DataFrame(columns=[
            "月份", "房间号码", "入住人员", "有效居住天数",
            "水电分摊金额(元)", "住宿费(元)", "总应付金额(元)"
        ])

    room_summary = []
    for rn, months in sorted(results_room.items()):
        for m, total in months.items():
            room_summary.append({
                "月份": m, "房间号码": rn,
                "房间总应付金额(元)": round(total, 2)
            })
    df_room = pd.DataFrame(room_summary)
    if len(df_room) == 0:
        df_room = pd.DataFrame(columns=["月份", "房间号码", "房间总应付金额(元)"])

    if progress_callback:
        progress_callback(98, 100, "正在保存...")

    base = os.path.splitext(os.path.basename(filepath))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.dirname(filepath) or os.getcwd()
    output_path = os.path.join(output_dir, f"分摊结果_{base}_{timestamp}.xlsx")
    _write_excel(output_path, df_person, df_room)

    if progress_callback:
        progress_callback(100, 100, "处理完成!")

    return {
        "output_path": output_path,
        "person_count": len(df_person),
        "room_count": len(df_room),
        "total_amount": df_room["房间总应付金额(元)"].sum()
        if len(df_room) > 0 and "房间总应付金额(元)" in df_room.columns else 0,
        "errors": errors,
        "billing_period": f"{billing_year}-{billing_month:02d}",
    }
