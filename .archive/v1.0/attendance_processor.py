# -*- coding: utf-8 -*-
import math
import os
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ====== Configurable settings ======
LATE_RULES = {
    "白": {"check_time": "08:00", "time_type": "first"},
    "晚": {"check_time": "20:00", "time_type": "last"},
}
LATE_MINUTES_FLOOR = True
SHOW_LATE_MINUTES_IN_CELL = True

# ====== Style constants ======
LATE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LATE_FONT = Font(name="Microsoft YaHei", size=9, color="9C0006")


def calc_late_minutes(actual_time, threshold_str):
    th = datetime.strptime(threshold_str, "%H:%M").time()
    if actual_time <= th:
        return 0
    dummy_date = datetime.min.date()
    dt_actual = datetime.combine(dummy_date, actual_time)
    dt_th = datetime.combine(dummy_date, th)
    delta_seconds = (dt_actual - dt_th).total_seconds()
    return int(delta_seconds // 60)


def get_donation(minutes):
    if 1 <= minutes <= 10:
        return (15, f"迟到{minutes}分钟")
    elif 11 <= minutes <= 30:
        return (30, f"迟到{minutes}分钟")
    elif 31 <= minutes <= 59:
        return (100, f"迟到{minutes}分钟")
    elif minutes >= 60:
        return (300, f"迟到{minutes}分钟")
    return (0, "")


class AttendanceProcessor:

    def __init__(self, exclude_names=None, progress_callback=None):
        self.exclude_names = set(exclude_names or [])
        self.progress_callback = progress_callback

    def _report_progress(self, current, total, message=""):
        if self.progress_callback:
            self.progress_callback(current, total, message)

    def parse_shift_schedule(self, shift_file_path):
        wb = openpyxl.load_workbook(shift_file_path, data_only=True)
        ws = wb.active
        date_map = {}
        for col in range(7, ws.max_column + 1):
            val = ws.cell(2, col).value
            if val is None:
                break
            if isinstance(val, (int, float)):
                d = datetime(1899, 12, 30) + timedelta(days=int(val))
                date_map[col] = d.date()
        shift_data = {}
        for row in range(4, ws.max_row + 1, 6):
            emp_name = ws.cell(row, 3).value
            if not emp_name:
                continue
            emp_name = str(emp_name).strip()
            for col, date in date_map.items():
                shift_val = ws.cell(row, col).value
                if shift_val:
                    shift_type = str(shift_val).strip()
                    if shift_type in LATE_RULES:
                        shift_data[(emp_name, date)] = shift_type
        wb.close()
        return shift_data

    @staticmethod
    def _check_late(punch_times, shift_type):
        if shift_type not in LATE_RULES:
            return (False, 0, None)
        rule = LATE_RULES[shift_type]
        time_type = rule["time_type"]
        actual = punch_times[0] if time_type == "first" else punch_times[1]
        minutes = calc_late_minutes(actual, rule["check_time"])
        if minutes >= 1:
            return (True, minutes, time_type)
        return (False, 0, time_type)

    def process(self, input_path, output_path, shift_file_path=None):
        self._report_progress(0, 100, "正在读取打卡文件...")
        wb = openpyxl.load_workbook(input_path)
        source_sheet_name = wb.sheetnames[0]
        for name in wb.sheetnames:
            if "考勤" in name or "打卡" in name:
                source_sheet_name = name
                break
        ws = wb[source_sheet_name]
        max_row = ws.max_row
        self._report_progress(5, 100,
            "找到工作表 {n}, 共 {c} 条记录".format(n=source_sheet_name, c=max_row - 1))

        shift_data = {}
        if shift_file_path and os.path.isfile(shift_file_path):
            self._report_progress(7, 100, "正在读取班次表...")
            shift_data = self.parse_shift_schedule(shift_file_path)
            self._report_progress(10, 100,
                "班次表: {n} 条记录".format(n=len(shift_data)))

        emp_data = defaultdict(lambda: defaultdict(list))
        all_dates = set()
        excluded_count = 0
        skipped_invalid = 0

        for row_idx in range(2, max_row + 1):
            if row_idx % 500 == 0:
                pct = 10 + int((row_idx - 2) / (max_row - 1) * 30)
                self._report_progress(pct, 100,
                    "Row {r}/{t}...".format(r=row_idx, t=max_row))
            date_val = ws.cell(row_idx, 1).value
            time_val = ws.cell(row_idx, 2).value
            name_val = ws.cell(row_idx, 3).value
            if not name_val:
                skipped_invalid += 1
                continue
            emp_name = str(name_val).strip()
            if emp_name in self.exclude_names:
                excluded_count += 1
                continue
            if not date_val:
                skipped_invalid += 1
                continue
            if isinstance(date_val, datetime):
                parsed_date = date_val.date()
            elif isinstance(date_val, str):
                try:
                    parsed_date = datetime.strptime(date_val.strip(), "%Y/%m/%d").date()
                except ValueError:
                    skipped_invalid += 1
                    continue
            else:
                skipped_invalid += 1
                continue
            if isinstance(time_val, datetime):
                parsed_time = time_val.time()
            elif isinstance(time_val, str):
                try:
                    parsed_time = datetime.strptime(time_val.strip(), "%H:%M:%S").time()
                except ValueError:
                    skipped_invalid += 1
                    continue
            else:
                skipped_invalid += 1
                continue
            all_dates.add(parsed_date)
            if not emp_data[emp_name][parsed_date]:
                emp_data[emp_name][parsed_date] = [parsed_time, parsed_time]
            else:
                if parsed_time < emp_data[emp_name][parsed_date][0]:
                    emp_data[emp_name][parsed_date][0] = parsed_time
                if parsed_time > emp_data[emp_name][parsed_date][1]:
                    emp_data[emp_name][parsed_date][1] = parsed_time

        self._report_progress(40, 100, "数据读取完成, 正在生成汇总表...")
        sorted_dates = sorted(all_dates)
        self._report_progress(50, 100,
            "{e} 名员工, {d} 个日期".format(e=len(emp_data), d=len(sorted_dates)))

        for sn in ["汇总考勤表", "迟到明细"]:
            if sn in wb.sheetnames:
                del wb[sn]
        ws_out = wb.create_sheet("汇总考勤表")

        hdr_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        count_align = Alignment(horizontal="center", vertical="center")
        name_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )
        even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        normal_font = Font(name="Microsoft YaHei", size=9)
        name_font = Font(name="Microsoft YaHei", size=10)

        total_cols = 1 + len(sorted_dates) + 1
        late_count_col = total_cols

        cell = ws_out.cell(1, 1, "姓名")
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border
        for ci, d in enumerate(sorted_dates, start=2):
            cell = ws_out.cell(1, ci, d.day)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border
        cell = ws_out.cell(1, late_count_col, "迟到总次数")
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = border

        sorted_employees = sorted(emp_data.keys())
        total = len(sorted_employees)
        late_count = 0
        late_records = []
        late_days_per_emp = defaultdict(set)

        for ri, emp in enumerate(sorted_employees, start=2):
            if ri % 10 == 0:
                pct = 50 + int((ri - 2) / total * 35)
                self._report_progress(pct, 100, "正在写入 {e}...".format(e=emp))
            nc = ws_out.cell(ri, 1, emp)
            nc.alignment = name_align; nc.border = border; nc.font = name_font
            if ri % 2 == 0:
                nc.fill = even_fill
            for ci, d in enumerate(sorted_dates, start=2):
                cell = ws_out.cell(ri, ci)
                cell.border = border
                cell.alignment = cell_align
                cell.font = normal_font
                is_late = False
                if d in emp_data[emp]:
                    times = emp_data[emp][d]
                    t0_str = times[0].strftime("%H:%M:%S")
                    t1_str = times[1].strftime("%H:%M:%S")
                    shift_key = (emp, d)
                    if shift_key in shift_data:
                        shift_type = shift_data[shift_key]
                        late_result = self._check_late(times, shift_type)
                        is_late, late_mins, time_type = late_result
                        if is_late and SHOW_LATE_MINUTES_IN_CELL:
                            if time_type == "first":
                                cell.value = "{t0}(迟到{m}分钟)\n{t1}".format(
                                    t0=t0_str, t1=t1_str, m=late_mins)
                            else:
                                cell.value = "{t0}\n{t1}(迟到{m}分钟)".format(
                                    t0=t0_str, t1=t1_str, m=late_mins)
                        else:
                            cell.value = "{t0}\n{t1}".format(t0=t0_str, t1=t1_str)
                        if is_late:
                            cell.fill = LATE_FILL
                            cell.font = LATE_FONT
                            late_count += 1
                            late_days_per_emp[emp].add(d)
                            late_records.append({
                                "employee": emp,
                                "date": d,
                                "shift_type": shift_type,
                                "threshold": LATE_RULES[shift_type]["check_time"],
                                "actual_time": times[0] if time_type == "first" else times[1],
                                "late_minutes": late_mins,
                            })
                    else:
                        cell.value = "{t0}\n{t1}".format(t0=t0_str, t1=t1_str)
                else:
                    cell.value = ""
                if not is_late and ri % 2 == 0:
                    cell.fill = even_fill
            tc = ws_out.cell(ri, late_count_col, len(late_days_per_emp.get(emp, set())))
            tc.border = border
            tc.alignment = count_align
            tc.font = normal_font
            if ri % 2 == 0:
                tc.fill = even_fill

        self._report_progress(85, 100, "正在调整格式...")
        ws_out.column_dimensions["A"].width = 10
        for ci in range(2, len(sorted_dates) + 2):
            ws_out.column_dimensions[get_column_letter(ci)].width = 13
        ws_out.column_dimensions[get_column_letter(late_count_col)].width = 11
        ws_out.freeze_panes = "B2"
        ws_out.auto_filter.ref = "A1:{col}{row}".format(
            col=get_column_letter(total_cols),
            row=len(sorted_employees) + 1)

        legend_row = len(sorted_employees) + 3
        if late_count > 0:
            lc = ws_out.cell(legend_row, 1, "图例:")
            lc.font = Font(name="Microsoft YaHei", size=9, bold=True)
            ll = ws_out.cell(legend_row, 2,
                "浅红底 = 迟到 (括号内为分钟数)")
            ll.fill = LATE_FILL
            ll.font = LATE_FONT
            ll.alignment = Alignment(horizontal="left", vertical="center")
            ws_out.merge_cells(start_row=legend_row, start_column=2,
                end_row=legend_row, end_column=min(6, total_cols))

        # Late detail sheet
        self._report_progress(90, 100, "正在生成迟到明细...")
        if "迟到明细" in wb.sheetnames:
            del wb["迟到明细"]
        ws_detail = wb.create_sheet("迟到明细")

        dh_headers = ["员工姓名", "日期", "班次",
                       "应到阈值", "实际打卡时间", "迟到分钟数",
                       "乐捐金额", "乐捐原因", "备注"]
        dh_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        dh_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        dh_align = Alignment(horizontal="center", vertical="center")
        for ci, h in enumerate(dh_headers, start=1):
            cell = ws_detail.cell(1, ci, h)
            cell.font = dh_font; cell.fill = dh_fill; cell.alignment = dh_align; cell.border = border

        late_records.sort(key=lambda x: (x["date"], x["employee"]))
        for ri, rec in enumerate(late_records, start=2):
            donation_amount, donation_reason = get_donation(rec["late_minutes"])
            vals = [
                rec["employee"],
                rec["date"].strftime("%Y-%m-%d"),
                rec["shift_type"],
                rec["threshold"],
                rec["actual_time"].strftime("%H:%M:%S"),
                rec["late_minutes"],
                donation_amount,
                donation_reason,
                ""
            ]
            for ci, v in enumerate(vals, start=1):
                cell = ws_detail.cell(ri, ci, v)
                cell.border = border
                cell.alignment = dh_align
                cell.font = Font(name="Microsoft YaHei", size=9)
                if ri % 2 == 0:
                    cell.fill = even_fill

        detail_widths = [10, 12, 6, 10, 12, 10, 10, 10, 10]
        for ci, w in enumerate(detail_widths, start=1):
            ws_detail.column_dimensions[get_column_letter(ci)].width = w
        ws_detail.freeze_panes = "A2"
        ws_detail.auto_filter.ref = "A1:I{row}".format(row=len(late_records) + 1)

        self._report_progress(98, 100, "正在保存...")
        wb.save(output_path)
        wb.close()
        self._report_progress(100, 100, "处理完成!")

        return {
            "total_records": max_row - 1,
            "valid_records": max_row - 1 - excluded_count - skipped_invalid,
            "excluded_count": excluded_count,
            "skipped_invalid": skipped_invalid,
            "employee_count": len(sorted_employees),
            "date_count": len(sorted_dates),
            "date_range": "{s} ~ {e}".format(s=sorted_dates[0], e=sorted_dates[-1]) if sorted_dates else "N/A",
            "output_path": output_path,
            "late_count": late_count,
            "shift_count": len(shift_data),
            "detail_records": len(late_records),
        }
